import openpyxl


class HomePageData:

    test_HomePage_Data = [{"firstname": "Srinivasan", "email": "test@gmail.com", "gender": "Male"},
                            {"firstname": "Srini", "email": "test@yahoo.com.com", "gender": "Female"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\ssakthivel\\Downloads\\TestCase.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    # print(f"Cell value of {i}, {j} is : ", sheet.cell(row=i, column=j).value)
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]
